from __future__ import annotations

from datetime import datetime
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd

from ngsatml.cohd import add_cohd_targets, parse_state_cohd_workbook
from ngsatml.cpi import add_cpi_targets, parse_state_food_cpi_workbook
from ngsatml.nbs import NIGERIA_STATES
from ngsatml.nbs_targets import audit_panel


def _bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cpi_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table-5"
    ws.cell(7, 1, "Combined Urban And Rural State Consumer Price Index")
    ws.cell(8, 2, "Base Period")
    ws.cell(9, 3, datetime(2024, 1, 1))
    ws.cell(9, 5, datetime(2024, 2, 1))
    ws.cell(9, 7, "Annual Change")
    ws.cell(10, 2, "State")
    ws.cell(10, 3, "Food")
    ws.cell(10, 4, "All Items")
    ws.cell(10, 5, "Food")
    ws.cell(10, 6, "All Items")
    ws.cell(10, 7, "Food")
    for i, state in enumerate(NIGERIA_STATES, start=12):
        ws.cell(i, 2, state)
        offset = i - 12
        ws.cell(i, 3, 700.0 + offset)
        ws.cell(i, 4, 600.0 + offset)
        ws.cell(i, 5, 714.0 + offset)
        ws.cell(i, 6, 610.0 + offset)
        ws.cell(i, 7, 2.0)
    return _bytes(wb)


def _cohd_workbook(month_value: float = 900.0) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CoHD by national average"
    ws.cell(1, 1, "CoHD State Average (Naira / person / day)")
    ws.cell(2, 1, "State")
    ws.cell(2, 2, "CoHD Average")
    for i, state in enumerate(NIGERIA_STATES, start=3):
        ws.cell(i, 1, "Abuja" if state == "FCT" else state)
        ws.cell(i, 2, month_value + i)
    return _bytes(wb)


def test_cpi_parser_accepts_excel_date_headers_and_rejects_change_columns():
    frame = parse_state_food_cpi_workbook(
        _cpi_workbook(),
        source_url="https://example.test/cpi.xlsx",
        workbook_name="cpi_February2024.xlsx",
        index_regime="2009-11-base",
    )
    assert set(frame["month"]) == {pd.Timestamp("2024-01-01"), pd.Timestamp("2024-02-01")}
    assert len(frame) == 74
    assert frame.groupby("month")["state"].nunique().eq(37).all()
    assert not np.isclose(frame["food_cpi"], 2.0).any()


def test_cpi_targets_use_exact_calendar_months_within_regime():
    frame = parse_state_food_cpi_workbook(
        _cpi_workbook(),
        source_url="https://example.test/cpi.xlsx",
        workbook_name="cpi_February2024.xlsx",
        index_regime="2009-11-base",
    )
    out = add_cpi_targets(frame, lags=(1,))
    feb = out[(out["state"] == "Abia") & (out["month"] == pd.Timestamp("2024-02-01"))].iloc[0]
    jan = out[(out["state"] == "Abia") & (out["month"] == pd.Timestamp("2024-01-01"))].iloc[0]
    assert feb["food_cpi_lag_1m"] == jan["food_cpi"]
    assert jan["target_food_cpi_1m"] == feb["food_cpi"]


def test_cohd_parser_and_targets_cover_all_37_states():
    jan = parse_state_cohd_workbook(
        _cohd_workbook(900),
        month=pd.Timestamp("2024-01-01"),
        source_url="https://example.test/cohd-jan.xlsx",
        workbook_name="January 2024_Table.xlsx",
    )
    feb = parse_state_cohd_workbook(
        _cohd_workbook(950),
        month=pd.Timestamp("2024-02-01"),
        source_url="https://example.test/cohd-feb.xlsx",
        workbook_name="February 2024_Table.xlsx",
    )
    assert len(jan) == 37
    assert "FCT" in set(jan["state"])
    out = add_cohd_targets(pd.concat([jan, feb], ignore_index=True), lags=(1,))
    feb_abia = out[(out["state"] == "Abia") & (out["month"] == pd.Timestamp("2024-02-01"))].iloc[0]
    jan_abia = out[(out["state"] == "Abia") & (out["month"] == pd.Timestamp("2024-01-01"))].iloc[0]
    assert feb_abia["cohd_lag_1m"] == jan_abia["cohd_ngn_person_day"]


def test_panel_audit_detects_complete_and_missing_months():
    rows = []
    for month in (pd.Timestamp("2024-01-01"), pd.Timestamp("2024-02-01")):
        for state in NIGERIA_STATES:
            rows.append({"state": state, "month": month, "value": 1.0})
    complete = pd.DataFrame(rows)
    assert audit_panel(complete, value_col="value")["ok"] is True

    missing_month = complete[complete["month"] == pd.Timestamp("2024-01-01")].copy()
    march = missing_month.copy()
    march["month"] = pd.Timestamp("2024-03-01")
    broken = pd.concat([missing_month, march], ignore_index=True)
    audit = audit_panel(broken, value_col="value")
    assert audit["ok"] is False
    assert audit["missing_months"][0]["months"] == ["2024-02"]
