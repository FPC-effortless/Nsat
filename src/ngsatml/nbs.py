from __future__ import annotations

import html as html_lib
import math
import re
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin

import numpy as np
import openpyxl
import pandas as pd
import requests

CATALOG_URL = "https://microdata.nigerianstat.gov.ng/index.php/catalog/162/related-materials"
ELIBRARY_URL = "https://www.nigerianstat.gov.ng/elibrary/"

NIGERIA_STATES = [
    "Abia", "Adamawa", "Akwa Ibom", "Anambra", "Bauchi", "Bayelsa", "Benue", "Borno",
    "Cross River", "Delta", "Ebonyi", "Edo", "Ekiti", "Enugu", "FCT", "Gombe", "Imo",
    "Jigawa", "Kaduna", "Kano", "Katsina", "Kebbi", "Kogi", "Kwara", "Lagos", "Nasarawa",
    "Niger", "Ogun", "Ondo", "Osun", "Oyo", "Plateau", "Rivers", "Sokoto", "Taraba",
    "Yobe", "Zamfara",
]

_STATE_ALIASES = {
    "abuja": "FCT",
    "fct": "FCT",
    "fct abuja": "FCT",
    "abuja fct": "FCT",
    "federal capital territory": "FCT",
    "akwa ibom": "Akwa Ibom",
    "akwaibom": "Akwa Ibom",
    "cross river": "Cross River",
    "crossriver": "Cross River",
    "nasarawa": "Nasarawa",
}
for _state in NIGERIA_STATES:
    _STATE_ALIASES.setdefault(_state.casefold(), _state)

_MONTHS = {
    "january": 1, "jan": 1,
    "february": 2, "feb": 2,
    "march": 3, "mar": 3,
    "april": 4, "apr": 4,
    "may": 5,
    "june": 6, "jun": 6,
    "july": 7, "jul": 7,
    "august": 8, "aug": 8,
    "september": 9, "sept": 9, "sep": 9,
    "october": 10, "oct": 10,
    "november": 11, "nov": 11,
    "december": 12, "dec": 12,
}

# Deliberately conservative mapping. Only entries with an unambiguous WFP analogue
# are used in cross-source comparisons.
_WFP_MAP = [
    (re.compile(r"rice.*local|local.*rice", re.I), "Rice (local)"),
    (re.compile(r"rice.*import", re.I), "Rice (imported)"),
    (re.compile(r"beans?.*brown", re.I), "Cowpeas (brown)"),
    (re.compile(r"maize.*white|white.*maize", re.I), "Maize (white)"),
    (re.compile(r"maize.*yellow|yellow.*maize", re.I), "Maize (yellow)"),
    (re.compile(r"\bmaize\b", re.I), "Maize"),
    (re.compile(r"\btomato", re.I), "Tomatoes"),
    (re.compile(r"\bonion", re.I), "Onions"),
    (re.compile(r"yam.*tuber|\byam\b", re.I), "Yam"),
    (re.compile(r"beef.*boneless|boneless.*beef", re.I), "Meat (beef)"),
    (re.compile(r"garri.*white|gari.*white", re.I), "Gari (white)"),
    (re.compile(r"garri.*yellow|gari.*yellow", re.I), "Cassava meal (gari, yellow)"),
    (re.compile(r"palm.*oil|oil.*palm", re.I), "Oil (palm)"),
    (re.compile(r"vegetable.*oil|oil.*vegetable", re.I), "Oil (vegetable)"),
    (re.compile(r"\begg", re.I), "Eggs"),
    (re.compile(r"milk.*powder|powder.*milk", re.I), "Milk (powder)"),
]


@dataclass(frozen=True)
class NBSResource:
    title: str
    url: str
    month: pd.Timestamp | None = None
    source_page: str | None = None
    source: str = "nbs"


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = html_lib.unescape(str(value))
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_state(value: object) -> str | None:
    text = _clean_text(value).casefold()
    text = re.sub(r"\bstate\b", "", text)
    text = re.sub(r"[^a-z ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return _STATE_ALIASES.get(text)


def month_from_text(text: str) -> pd.Timestamp | None:
    cleaned = _clean_text(text).casefold().replace("_", " ").replace("-", " ")
    year_match = re.search(r"\b(20\d{2})\b", cleaned)
    if not year_match:
        return None
    year = int(year_match.group(1))
    tokens = re.findall(r"[a-z]+", cleaned)
    for token in tokens:
        if token in _MONTHS:
            return pd.Timestamp(year=year, month=_MONTHS[token], day=1)
    return None


def canonical_wfp_commodity(item: str) -> str | None:
    for pattern, value in _WFP_MAP:
        if pattern.search(item):
            return value
    return None


def parse_item_unit(item: str) -> tuple[str, float, str]:
    """Return canonical unit, package quantity in that unit, and family.

    NBS item labels commonly encode units in the item name (1kg, 500g,
    12 pieces, one bottle). Unknown package definitions are retained as
    package-level values and are excluded from cross-source unit comparisons.
    """
    text = item.casefold().replace(",", " ")
    match = re.search(r"(\d+(?:\.\d+)?)\s*(kg|kilograms?|g|grams?|litres?|liters?|ltr|l\b|pieces?|pcs\b|tubers?)", text)
    if match:
        qty = float(match.group(1))
        unit = match.group(2)
        if unit in {"kg", "kilogram", "kilograms"}:
            return "kg", qty, "mass"
        if unit in {"g", "gram", "grams"}:
            return "kg", qty / 1000.0, "mass"
        if unit in {"litre", "litres", "liter", "liters", "ltr", "l"}:
            return "l", qty, "volume"
        return "item", qty, "count"
    if re.search(r"\bone\s+bottle\b|\b1\s*bottle\b", text):
        return "bottle", 1.0, "package"
    if "bottle" in text:
        return "bottle", 1.0, "package"
    return "package", 1.0, "package"


def discover_downloads(session: requests.Session | None = None) -> list[NBSResource]:
    """Discover NBS NADA resources, best-effort.

    The NADA host is occasionally slow from GitHub runners, so production
    builds should prefer direct e-library table links and use this as a recent
    resource supplement rather than the sole source.
    """
    s = session or requests.Session()
    r = s.get(CATALOG_URL, timeout=(15, 30))
    r.raise_for_status()
    html = r.text
    found: list[NBSResource] = []
    pattern = re.compile(r'href=["\']([^"\']*/catalog/162/download/\d+)["\']', re.I)
    for m in pattern.finditer(html):
        url = urljoin(CATALOG_URL, m.group(1))
        context = _clean_text(html[max(0, m.start() - 700):m.start() + 100])
        month = month_from_text(context)
        found.append(NBSResource(title=context[-240:] or "NBS resource", url=url, month=month, source_page=CATALOG_URL, source="nada"))
    dedup: dict[str, NBSResource] = {}
    for item in found:
        dedup.setdefault(item.url, item)
    return list(dedup.values())


def discover_elibrary_table_downloads(
    *,
    start_date: str | pd.Timestamp = "2022-01-01",
    end_date: str | pd.Timestamp | None = None,
    session: requests.Session | None = None,
) -> list[NBSResource]:
    """Discover direct XLS/XLSX table files from archived NBS report pages."""
    s = session or requests.Session()
    s.headers.setdefault("User-Agent", "Nsat/0.3 (+https://github.com/FPC-effortless/Nsat)")
    response = s.get(ELIBRARY_URL, timeout=(20, 60))
    response.raise_for_status()
    body = response.text
    start = pd.Timestamp(start_date)
    end = pd.Timestamp(end_date) if end_date is not None else pd.Timestamp.utcnow().tz_localize(None) + pd.offsets.MonthBegin(1)

    candidates: dict[str, tuple[str, pd.Timestamp]] = {}
    anchor_pattern = re.compile(
        r'<a[^>]+href=["\']([^"\']*/elibrary/read/\d+)["\'][^>]*>(.*?)</a>',
        re.I | re.S,
    )
    for href, label_html in anchor_pattern.findall(body):
        label = _clean_text(label_html)
        if "selected food prices watch" not in label.casefold():
            continue
        month = month_from_text(label)
        if month is None or not (start <= month < end):
            continue
        candidates[urljoin(ELIBRARY_URL, href)] = (label, month)

    # Some versions of the e-library render the report label outside the anchor.
    if not candidates:
        fallback = re.compile(
            r'href=["\']([^"\']*/elibrary/read/\d+)["\'][^>]*>.*?</a>.{0,300}?Selected Food Prices Watch\s*\(([^)]+)\)',
            re.I | re.S,
        )
        for href, label_part in fallback.findall(body):
            label = f"Selected Food Prices Watch ({_clean_text(label_part)})"
            month = month_from_text(label)
            if month is not None and start <= month < end:
                candidates[urljoin(ELIBRARY_URL, href)] = (label, month)

    resources: list[NBSResource] = []
    for page_url, (title, month) in sorted(candidates.items(), key=lambda x: x[1][1]):
        try:
            page = s.get(page_url, timeout=(15, 35))
            page.raise_for_status()
        except requests.RequestException:
            continue
        links = re.findall(r'href=["\']([^"\']+\.(?:xlsx|xls)(?:\?[^"\']*)?)["\']', page.text, re.I)
        if not links:
            continue
        # Prefer a link whose nearby context says tables; report pages generally
        # expose one table workbook and one PDF report.
        url = urljoin(page_url, links[0])
        resources.append(NBSResource(title=title, url=url, month=month, source_page=page_url, source="elibrary"))
    return resources


def workbook_payloads(content: bytes, *, fallback_name: str = "download.xlsx") -> list[tuple[str, bytes]]:
    if content[:2] == b"PK":
        try:
            with zipfile.ZipFile(BytesIO(content)) as zf:
                files = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xlsm"))]
                if files:
                    return [(Path(name).name, zf.read(name)) for name in files]
        except zipfile.BadZipFile:
            pass
    return [(fallback_name, content)]


def download_resource(
    resource: NBSResource,
    cache_dir: str | Path,
    *,
    session: requests.Session | None = None,
) -> list[Path]:
    cache = Path(cache_dir)
    cache.mkdir(parents=True, exist_ok=True)
    s = session or requests.Session()
    s.headers.setdefault("User-Agent", "Nsat/0.3 (+https://github.com/FPC-effortless/Nsat)")
    response = s.get(resource.url, timeout=(20, 60))
    response.raise_for_status()
    fallback = Path(resource.url.split("?", 1)[0]).name or "download.xlsx"
    paths: list[Path] = []
    for idx, (name, payload) in enumerate(workbook_payloads(response.content, fallback_name=fallback)):
        month = month_from_text(name) or resource.month
        prefix = month.strftime("%Y-%m") if month is not None else "unknown-month"
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
        target = cache / f"{prefix}_{idx:02d}_{safe_name}"
        target.write_bytes(payload)
        paths.append(target)
    return paths


def _to_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.number)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = _clean_text(value).replace(",", "").replace("₦", "").replace("N", "")
    text = text.replace("-", "") if re.fullmatch(r"\s*-\s*", str(value)) else text
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        number = float(match.group(0))
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _item_label(rows: list[list[object]], first_state_row: int, col: int, sheet: str) -> str:
    pieces: list[str] = []
    start = max(0, first_state_row - 6)
    for r in range(start, first_state_row):
        if col >= len(rows[r]):
            continue
        text = _clean_text(rows[r][col])
        if not text:
            continue
        if text.casefold() in {"state", "states", "s/n", "sn", "average price", "avg price", "price"}:
            continue
        if text not in pieces:
            pieces.append(text)
    if not pieces and sheet:
        pieces = [_clean_text(sheet)]
    return " | ".join(pieces).strip(" |")


def parse_state_price_workbook(
    payload: bytes,
    *,
    month: pd.Timestamp,
    source_url: str,
    workbook_name: str = "workbook.xlsx",
) -> pd.DataFrame:
    """Extract state-level price cells without assuming a fixed NBS sheet layout.

    The parser finds columns containing Nigerian state names, then melts numeric
    columns across those state rows. It therefore tolerates title/header rows and
    multi-sheet workbooks that changed format over time.
    """
    wb = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    records: list[dict[str, object]] = []

    for ws in wb.worksheets:
        rows: list[list[object]] = []
        max_cols = min(int(ws.max_column or 0), 120)
        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=0):
            rows.append(list(row[:max_cols]))
            if ridx >= 180:
                break
        if not rows or max_cols == 0:
            continue

        best_col = None
        best_matches: list[tuple[int, str]] = []
        for col in range(max_cols):
            matches: list[tuple[int, str]] = []
            for ridx, row in enumerate(rows):
                if col >= len(row):
                    continue
                state = normalize_state(row[col])
                if state is not None:
                    matches.append((ridx, state))
            unique_states = {state for _, state in matches}
            if len(unique_states) > len({s for _, s in best_matches}):
                best_col = col
                best_matches = matches

        unique_states = {s for _, s in best_matches}
        if best_col is None or len(unique_states) < 15:
            continue
        state_rows = sorted({ridx for ridx, _ in best_matches})
        first_state_row = min(state_rows)

        for col in range(max_cols):
            if col == best_col:
                continue
            label = _item_label(rows, first_state_row, col, ws.title)
            label_low = label.casefold()
            if not label or any(token in label_low for token in ["% change", "percent change", "percentage change", "rank", "position"]):
                continue
            values: list[tuple[str, float]] = []
            for ridx, state in best_matches:
                if ridx >= len(rows) or col >= len(rows[ridx]):
                    continue
                number = _to_number(rows[ridx][col])
                if number is not None and number > 0:
                    values.append((state, number))
            if len({state for state, _ in values}) < 10:
                continue

            unit, quantity, family = parse_item_unit(label)
            mapped = canonical_wfp_commodity(label)
            for state, price in values:
                records.append({
                    "state": state,
                    "month": pd.Timestamp(month).to_period("M").to_timestamp(),
                    "item": label,
                    "wfp_commodity": mapped,
                    "unit": unit,
                    "unit_family": family,
                    "unit_quantity": quantity,
                    "price_ngn": float(price),
                    "price_ngn_base": float(price / quantity) if quantity > 0 else np.nan,
                    "source_sheet": ws.title,
                    "source_workbook": workbook_name,
                    "source_url": source_url,
                })

    if not records:
        return pd.DataFrame(columns=[
            "state", "month", "item", "wfp_commodity", "unit", "unit_family",
            "unit_quantity", "price_ngn", "price_ngn_base", "source_sheet",
            "source_workbook", "source_url",
        ])
    out = pd.DataFrame(records)
    out = out.drop_duplicates(["state", "month", "item", "source_sheet", "source_workbook"])
    return out.sort_values(["month", "state", "item"]).reset_index(drop=True)


def parse_workbook_file(path: str | Path, *, month: pd.Timestamp, source_url: str) -> pd.DataFrame:
    p = Path(path)
    return parse_state_price_workbook(p.read_bytes(), month=month, source_url=source_url, workbook_name=p.name)


def consolidate_state_prices(frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    parts = [f for f in frames if f is not None and not f.empty]
    if not parts:
        return pd.DataFrame()
    data = pd.concat(parts, ignore_index=True)
    key = ["state", "month", "item", "unit"]
    agg = {
        "wfp_commodity": "first",
        "unit_family": "first",
        "unit_quantity": "first",
        "price_ngn": "median",
        "price_ngn_base": "median",
        "source_sheet": lambda s: "|".join(sorted(set(map(str, s)))),
        "source_workbook": lambda s: "|".join(sorted(set(map(str, s)))),
        "source_url": lambda s: "|".join(sorted(set(map(str, s)))),
    }
    return data.groupby(key, dropna=False, as_index=False).agg(agg).sort_values(key).reset_index(drop=True)


def add_state_calendar_targets(
    frame: pd.DataFrame,
    *,
    lags: tuple[int, ...] = (1, 2, 3, 6, 12),
) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    keys = ["state", "item", "unit"]
    out = frame.copy()
    lookup = out[keys + ["month", "price_ngn_base"]].drop_duplicates(keys + ["month"])
    for lag in lags:
        prior = lookup.copy()
        prior["month"] = prior["month"] + pd.DateOffset(months=lag)
        prior = prior.rename(columns={"price_ngn_base": f"price_lag_{lag}m"})
        out = out.merge(prior, on=keys + ["month"], how="left")
    future = lookup.copy()
    future["month"] = future["month"] - pd.DateOffset(months=1)
    future = future.rename(columns={"price_ngn_base": "target_price_ngn_1m"})
    out = out.merge(future, on=keys + ["month"], how="left")
    out["target_month"] = out["month"] + pd.DateOffset(months=1)
    out["target_log_change_1m"] = np.where(
        (out["price_ngn_base"] > 0) & (out["target_price_ngn_1m"] > 0),
        np.log(out["target_price_ngn_1m"] / out["price_ngn_base"]),
        np.nan,
    )
    out["target_change_1m_pct"] = np.where(
        out["price_ngn_base"] > 0,
        (out["target_price_ngn_1m"] - out["price_ngn_base"]) / out["price_ngn_base"],
        np.nan,
    )
    out["year"] = out["month"].dt.year.astype("int16")
    out["month_number"] = out["month"].dt.month.astype("int8")
    angle = 2 * math.pi * (out["month_number"].astype(float) - 1) / 12
    out["month_sin"] = np.sin(angle)
    out["month_cos"] = np.cos(angle)
    lag_cols = [f"price_lag_{lag}m" for lag in lags]
    out["lag_feature_count"] = out[lag_cols].notna().sum(axis=1).astype("int8")
    out["quality_suspicious"] = out["target_log_change_1m"].abs().gt(math.log(5)).fillna(False)
    return out


def aggregate_wfp_state_month(wfp: pd.DataFrame) -> pd.DataFrame:
    required = {"admin1", "month", "commodity", "unit", "price_ngn"}
    missing = required - set(wfp.columns)
    if missing:
        raise ValueError(f"WFP frame missing columns: {sorted(missing)}")
    work = wfp.copy()
    work["state"] = work["admin1"].map(normalize_state)
    work = work[work["state"].notna()]
    return (
        work.groupby(["state", "month", "commodity", "unit"], as_index=False)
        .agg(
            wfp_price_ngn=("price_ngn", "median"),
            wfp_market_rows=("market_id", "size") if "market_id" in work.columns else ("price_ngn", "size"),
            wfp_markets=("market_id", "nunique") if "market_id" in work.columns else ("price_ngn", "size"),
        )
    )


def build_wfp_overlap(nbs: pd.DataFrame, wfp: pd.DataFrame) -> pd.DataFrame:
    if nbs.empty or wfp.empty:
        return pd.DataFrame()
    n = nbs[nbs["wfp_commodity"].notna() & nbs["unit"].isin(["kg", "l", "item"])].copy()
    n = n.rename(columns={"price_ngn_base": "nbs_price_ngn"})
    w = aggregate_wfp_state_month(wfp)
    merged = n.merge(
        w,
        left_on=["state", "month", "wfp_commodity", "unit"],
        right_on=["state", "month", "commodity", "unit"],
        how="inner",
    )
    merged["nbs_wfp_ratio"] = merged["nbs_price_ngn"] / merged["wfp_price_ngn"]
    merged["nbs_wfp_log_ratio"] = np.log(merged["nbs_wfp_ratio"])
    merged["nbs_wfp_abs_pct_diff"] = (merged["nbs_price_ngn"] - merged["wfp_price_ngn"]).abs() / merged["wfp_price_ngn"]
    return merged.sort_values(["month", "state", "wfp_commodity"]).reset_index(drop=True)


def save_resource_index(resources: list[NBSResource], path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {
            "title": r.title,
            "url": r.url,
            "month": r.month.strftime("%Y-%m-%d") if r.month is not None else None,
            "source_page": r.source_page,
            "source": r.source,
        }
        for r in resources
    ]).to_csv(p, index=False)
    return p
