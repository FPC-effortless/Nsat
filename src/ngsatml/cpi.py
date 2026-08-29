from __future__ import annotations

import math
import re
from datetime import date, datetime
from io import BytesIO
from typing import Iterable, Iterator

import numpy as np
import openpyxl
import pandas as pd
import xlrd

from .nbs import normalize_state

_MONTH_RE = re.compile(
    r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(20\d{2})\b",
    re.I,
)
_MONTH_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}
_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _month(value: object) -> pd.Timestamp | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        ts = pd.Timestamp(value)
        return pd.Timestamp(year=ts.year, month=ts.month, day=1)
    if isinstance(value, (int, float, np.number)):
        return None
    text = str(value).strip()
    m = _MONTH_RE.search(text)
    if m:
        return pd.Timestamp(year=int(m.group(2)), month=_MONTH_NUM[m.group(1).casefold()], day=1)
    # NBS legacy workbooks frequently store month headers as Excel dates, which
    # openpyxl exposes as datetime objects but schema probes may stringify.
    try:
        ts = pd.to_datetime(text, errors="raise")
    except (ValueError, TypeError, OverflowError):
        return None
    if pd.isna(ts) or not (2000 <= int(ts.year) <= 2100):
        return None
    return pd.Timestamp(year=int(ts.year), month=int(ts.month), day=1)


def _num(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.number)):
        x = float(value)
        return x if math.isfinite(x) else None
    text = str(value).replace(",", "").strip()
    try:
        x = float(text)
    except ValueError:
        return None
    return x if math.isfinite(x) else None


def _sheet_rows_openxml(payload: bytes, *, max_rows: int, max_cols: int) -> Iterator[tuple[str, list[list[object]]]]:
    wb = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows: list[list[object]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row[:max_cols]))
            if i + 1 >= max_rows:
                break
        yield ws.title, rows


def _sheet_rows_xls(payload: bytes, *, max_rows: int, max_cols: int) -> Iterator[tuple[str, list[list[object]]]]:
    wb = xlrd.open_workbook(file_contents=payload, on_demand=True)
    for ws in wb.sheets():
        rows: list[list[object]] = []
        for ridx in range(min(ws.nrows, max_rows)):
            row: list[object] = []
            for cidx in range(min(ws.ncols, max_cols)):
                cell = ws.cell(ridx, cidx)
                value: object = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode)
                    except (ValueError, OverflowError):
                        pass
                row.append(value)
            rows.append(row)
        yield ws.name, rows


def _workbook_sheets(payload: bytes, *, max_rows: int = 220, max_cols: int = 80) -> list[tuple[str, list[list[object]]]]:
    if payload.startswith(_OLE2):
        return list(_sheet_rows_xls(payload, max_rows=max_rows, max_cols=max_cols))
    return list(_sheet_rows_openxml(payload, max_rows=max_rows, max_cols=max_cols))


def parse_state_food_cpi_workbook(
    payload: bytes,
    *,
    source_url: str,
    workbook_name: str,
    index_regime: str,
) -> pd.DataFrame:
    """Parse all-state Food CPI values from an official NBS workbook.

    The parser does not depend on a fixed worksheet name. It identifies the
    worksheet/column containing at least 30 canonical Nigerian states, recovers
    rolling month headers (including real Excel dates), and selects raw FOOD
    index columns while rejecting annual/monthly percentage-change columns.
    Both OpenXML and legacy binary XLS workbooks are supported.
    """
    candidates = sorted(
        _workbook_sheets(payload),
        key=lambda item: ("state cpi" not in item[0].casefold(), item[0].casefold()),
    )
    records: list[dict[str, object]] = []

    for sheet_name, rows in candidates:
        if not rows:
            continue
        width = max((len(r) for r in rows), default=0)
        best_state_col: int | None = None
        best_state_rows: list[tuple[int, str]] = []
        for col in range(width):
            found: list[tuple[int, str]] = []
            for ridx, row in enumerate(rows):
                if col < len(row):
                    state = normalize_state(row[col])
                    if state is not None:
                        found.append((ridx, state))
            if len({s for _, s in found}) > len({s for _, s in best_state_rows}):
                best_state_col, best_state_rows = col, found
        if best_state_col is None or len({s for _, s in best_state_rows}) < 30:
            continue

        first_state = min(r for r, _ in best_state_rows)
        header_rows = rows[max(0, first_state - 10):first_state]
        months_by_col: dict[int, pd.Timestamp] = {}
        last_month: pd.Timestamp | None = None
        for col in range(width):
            direct: pd.Timestamp | None = None
            for header in header_rows:
                if col < len(header):
                    parsed = _month(header[col])
                    if parsed is not None:
                        direct = parsed
            if direct is not None:
                last_month = direct
            if last_month is not None:
                months_by_col[col] = last_month

        for col, month in months_by_col.items():
            header_stack = " ".join(
                str(header[col]) for header in header_rows
                if col < len(header) and header[col] not in (None, "")
            ).casefold()
            if "food" not in header_stack:
                continue
            if any(term in header_stack for term in (
                "annual change", "monthly change", "% change", "percent change",
                "year on year", "month on month", "y-o-y", "m-o-m",
            )):
                continue
            values: list[tuple[str, float]] = []
            for ridx, state in best_state_rows:
                if ridx < len(rows) and col < len(rows[ridx]):
                    x = _num(rows[ridx][col])
                    if x is not None and x > 0:
                        values.append((state, x))
            if len({s for s, _ in values}) < 30:
                continue
            for state, food_index in values:
                records.append({
                    "state": state,
                    "month": month,
                    "food_cpi": food_index,
                    "index_regime": index_regime,
                    "source_sheet": sheet_name,
                    "source_workbook": workbook_name,
                    "source_url": source_url,
                })
        if records:
            break

    columns = [
        "state", "month", "food_cpi", "index_regime", "source_sheet",
        "source_workbook", "source_url",
    ]
    if not records:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame(records)
    return (
        out.groupby(["state", "month", "index_regime"], as_index=False)
        .agg(
            food_cpi=("food_cpi", "median"),
            source_sheet=("source_sheet", "first"),
            source_workbook=("source_workbook", "first"),
            source_url=("source_url", "first"),
        )
        .sort_values(["month", "state"])
        .reset_index(drop=True)
    )


def consolidate_cpi(frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    parts = [f for f in frames if f is not None and not f.empty]
    if not parts:
        return pd.DataFrame()
    all_rows = pd.concat(parts, ignore_index=True)
    grouped = (
        all_rows.groupby(["state", "month", "index_regime"], as_index=False)
        .agg(
            food_cpi=("food_cpi", "median"),
            source_count=("food_cpi", "size"),
            source_min=("food_cpi", "min"),
            source_max=("food_cpi", "max"),
            source_workbooks=("source_workbook", lambda s: "|".join(sorted(set(map(str, s))))),
            source_urls=("source_url", lambda s: "|".join(sorted(set(map(str, s))))),
        )
    )
    grouped["source_disagreement"] = (grouped["source_max"] - grouped["source_min"]).abs()
    return grouped.sort_values(["month", "state", "index_regime"]).reset_index(drop=True)


def add_cpi_targets(frame: pd.DataFrame, *, lags: tuple[int, ...] = (1, 2, 3, 6, 12)) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    keys = ["state", "index_regime"]
    out = frame.copy()
    lookup = out[keys + ["month", "food_cpi"]].drop_duplicates(keys + ["month"])
    for lag in lags:
        prior = lookup.copy()
        prior["month"] = prior["month"] + pd.DateOffset(months=lag)
        prior = prior.rename(columns={"food_cpi": f"food_cpi_lag_{lag}m"})
        out = out.merge(prior, on=keys + ["month"], how="left")
    future = lookup.copy()
    future["month"] = future["month"] - pd.DateOffset(months=1)
    future = future.rename(columns={"food_cpi": "target_food_cpi_1m"})
    out = out.merge(future, on=keys + ["month"], how="left")
    out["target_month"] = out["month"] + pd.DateOffset(months=1)
    out["target_food_cpi_log_change_1m"] = np.where(
        (out["food_cpi"] > 0) & (out["target_food_cpi_1m"] > 0),
        np.log(out["target_food_cpi_1m"] / out["food_cpi"]),
        np.nan,
    )
    out["target_food_cpi_change_1m_pct"] = np.where(
        out["food_cpi"] > 0,
        (out["target_food_cpi_1m"] - out["food_cpi"]) / out["food_cpi"],
        np.nan,
    )
    out["year"] = out["month"].dt.year.astype("int16")
    out["month_number"] = out["month"].dt.month.astype("int8")
    angle = 2.0 * math.pi * (out["month_number"].astype(float) - 1.0) / 12.0
    out["month_sin"] = np.sin(angle)
    out["month_cos"] = np.cos(angle)
    lag_cols = [f"food_cpi_lag_{lag}m" for lag in lags]
    out["lag_feature_count"] = out[lag_cols].notna().sum(axis=1).astype("int8")
    return out
