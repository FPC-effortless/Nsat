from __future__ import annotations

import json
import sys
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
import requests

# Official NBS e-library and NADA resources used only for schema regression probes.
# Production discovery is implemented in ngsatml.nbs_targets and must not depend
# on this fixed list.
RESOURCES = {
    "food_dec_2023": "https://nigerianstat.gov.ng/resource/selected_food_Dec_2023.xlsx",
    "food_jan_2024": "https://nigerianstat.gov.ng/resource/SELECTED_FOOD_JANUARY_2024.xlsx",
    "food_mar_2024": "https://nigerianstat.gov.ng/resource/selected_food_march_2024.xlsx",
    "food_oct_2024": "https://nigerianstat.gov.ng/resource/selected_food_oct_2024.xlsx",
    "cpi_jan_2024": "https://nigerianstat.gov.ng/resource/cpi_1NewJANUARY2024.xlsx",
    "cpi_jul_2024": "https://nigerianstat.gov.ng/resource/cpi_1NewJuly_2024.xlsx",
    "cpi_aug_2024": "https://nigerianstat.gov.ng/resource/cpi_1NewAug_2024.xlsx",
    "cpi_oct_2024": "https://nigerianstat.gov.ng/resource/cpi_OCT2024.xlsx",
    "cohd_jan_2024": "https://nigerianstat.gov.ng/resource/COHD_January_2024_Table.xlsx",
    "cohd_sep_2024": "https://nigerianstat.gov.ng/resource/cohd_sept2024.xlsx",
    "cpi_jan_2026": "https://microdata.nigerianstat.gov.ng/index.php/catalog/154/download/1353",
    "cpi_jul_2026": "https://microdata.nigerianstat.gov.ng/index.php/catalog/154/download/1432",
    "cohd_jan_feb_2026": "https://microdata.nigerianstat.gov.ng/index.php/catalog/146/download/1407",
    "cohd_apr_2026": "https://microdata.nigerianstat.gov.ng/index.php/catalog/146/download/1428",
}


def workbook_payloads(content: bytes, content_type: str) -> list[tuple[str, bytes]]:
    if content[:2] == b"PK":
        try:
            with zipfile.ZipFile(BytesIO(content)) as zf:
                names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xlsm"))]
                if names:
                    return [(Path(n).name, zf.read(n)) for n in names]
        except zipfile.BadZipFile:
            pass
    return [("download.xlsx", content)]


def inspect_book(name: str, payload: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    result = {"name": name, "sheets": []}
    for ws in wb.worksheets:
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [None if v is None else str(v) for v in row[:32]]
            if any(v not in (None, "") for v in vals):
                rows.append({"row": idx, "values": vals})
            if len(rows) >= 42:
                break
        result["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample": rows,
        })
    return result


def main(out_path: str) -> None:
    session = requests.Session()
    session.headers.update({"User-Agent": "Nsat/0.4 (+https://github.com/FPC-effortless/Nsat)"})
    report = {}
    for key, url in RESOURCES.items():
        try:
            response = session.get(url, timeout=(20, 60))
            response.raise_for_status()
            payloads = workbook_payloads(response.content, response.headers.get("content-type", ""))
            report[key] = {
                "url": url,
                "status": response.status_code,
                "content_type": response.headers.get("content-type"),
                "bytes": len(response.content),
                "workbooks": [inspect_book(name, data) for name, data in payloads],
            }
        except Exception as exc:
            report[key] = {"url": url, "error": f"{type(exc).__name__}: {exc}"}
        print(json.dumps({key: report[key]}, indent=2))
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    Path(out_path).write_text(json.dumps(report, indent=2), encoding="utf-8")
    if not any(v.get("workbooks") for v in report.values()):
        raise SystemExit("No NBS workbooks could be inspected")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "data/nbs/schema_report.json")
